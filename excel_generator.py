import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
import logging
import tempfile

logger = logging.getLogger(__name__)

def generate_excel_report(data, output_path=None):
    """
    Generate a formatted Excel report from the provided data
    
    Args:
        data: List of dictionaries containing the market data
        output_path: Path to save the Excel file (optional)
        
    Returns:
        str: Path to the generated Excel file
    """
    try:
        if not data:
            logger.error("No data available for Excel report")
            return None
            
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Drop MongoDB _id column if present
        if "_id" in df.columns:
            df = df.drop("_id", axis=1)
            
        # Sort by date
        df = df.sort_values("fecha")
        
        # Format date column
        df["fecha"] = pd.to_datetime(df["fecha"])
        
        # Generate output path if not provided
        if not output_path:
            # Check if running in Render (cloud environment)
            if os.environ.get("RENDER"):
                # Use a temporary directory that Render can write to
                temp_dir = tempfile.gettempdir()
                output_path = os.path.join(temp_dir, f"GeoPark_Report_{datetime.now().strftime('%Y%m%d')}.xlsx")
                logger.info(f"Running in Render environment, using temp path: {output_path}")
            else:
                output_path = f"GeoPark_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
        # Save to Excel
        df.to_excel(output_path, index=False)
        
        # Apply formatting
        apply_excel_formatting(output_path, df)
        
        logger.info(f"Excel report generated: {output_path}")
        return output_path
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        return None

def apply_excel_formatting(file_path, df):
    """Apply formatting to the Excel file"""
    try:
        # Load workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Define styles
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header styles
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Auto adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Apply data formatting
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                # Format numbers
                if isinstance(cell.value, (int, float)) and cell.column_letter != 'A':  # Skip date column
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                # Format dates
                elif cell.column_letter == 'A':  # Date column
                    cell.number_format = 'yyyy-mm-dd'
                    cell.alignment = Alignment(horizontal='center')
        
        # Check if running in Render (cloud environment)
        if not os.environ.get("RENDER"):
            # Add charts only in local environment
            add_price_chart(wb, ws, df)
            add_volume_chart(wb, ws, df)
        
        # Save the workbook
        wb.save(file_path)
        
        logger.info(f"Excel formatting applied to {file_path}")
        return True
    except Exception as e:
        logger.error(f"Error applying Excel formatting: {e}")
        return False

def add_price_chart(wb, ws, df):
    """Add price chart to the Excel workbook"""
    try:
        # Create a new sheet for the chart
        chart_sheet = wb.create_sheet(title="Price Chart")
        
        # Create chart
        chart = LineChart()
        chart.title = "GeoPark Price vs Brent"
        chart.style = 2
        chart.x_axis.title = "Date"
        chart.y_axis.title = "Price"
        
        # Get data ranges
        data_rows = len(df) + 1  # +1 for header
        
        # Add GeoPark price data
        price_col = df.columns.get_loc("precio_geo") + 1  # +1 because Excel is 1-indexed
        price_data = Reference(ws, min_col=price_col, min_row=1, max_row=data_rows)
        chart.add_data(price_data, titles_from_data=True)
        
        # Add Brent price data if available
        if "brent" in df.columns:
            brent_col = df.columns.get_loc("brent") + 1
            brent_data = Reference(ws, min_col=brent_col, min_row=1, max_row=data_rows)
            chart.add_data(brent_data, titles_from_data=True)
        
        # Add dates for X-axis
        dates = Reference(ws, min_col=1, min_row=2, max_row=data_rows)
        chart.set_categories(dates)
        
        # Add chart to the sheet
        chart_sheet.add_chart(chart, "B5")
        
        logger.info("Price chart added to Excel report")
        return True
    except Exception as e:
        logger.error(f"Error adding price chart: {e}")
        return False

def add_volume_chart(wb, ws, df):
    """Add volume chart to the Excel workbook"""
    try:
        # Get or create chart sheet
        if "Volume Chart" in wb.sheetnames:
            chart_sheet = wb["Volume Chart"]
        else:
            chart_sheet = wb.create_sheet(title="Volume Chart")
        
        # Create chart
        chart = LineChart()
        chart.title = "GeoPark Trading Volume"
        chart.style = 3
        chart.x_axis.title = "Date"
        chart.y_axis.title = "Volume"
        
        # Get data ranges
        data_rows = len(df) + 1  # +1 for header
        
        # Add volume data
        volume_col = df.columns.get_loc("volumen") + 1
        volume_data = Reference(ws, min_col=volume_col, min_row=1, max_row=data_rows)
        chart.add_data(volume_data, titles_from_data=True)
        
        # Add dates for X-axis
        dates = Reference(ws, min_col=1, min_row=2, max_row=data_rows)
        chart.set_categories(dates)
        
        # Add chart to the sheet
        chart_sheet.add_chart(chart, "B5")
        
        logger.info("Volume chart added to Excel report")
        return True
    except Exception as e:
        logger.error(f"Error adding volume chart: {e}")
        return False

if __name__ == "__main__":
    # Test with sample data
    sample_data = [
        {"fecha": "2023-10-01", "precio_geo": 10.25, "volumen": 150000, "brent": 85.75},
        {"fecha": "2023-10-02", "precio_geo": 10.50, "volumen": 160000, "brent": 86.25},
        {"fecha": "2023-10-03", "precio_geo": 10.35, "volumen": 145000, "brent": 85.50},
        {"fecha": "2023-10-04", "precio_geo": 10.60, "volumen": 170000, "brent": 87.00},
        {"fecha": "2023-10-05", "precio_geo": 10.75, "volumen": 180000, "brent": 87.50}
    ]
    
    generate_excel_report(sample_data, "Sample_Report.xlsx") 